"""Tests de API y servicio para Libro de Clases Digital."""

import json
from io import BytesIO
from datetime import date
from unittest.mock import patch

from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from backend.apps.accounts.models import Role, User
from backend.apps.academico.models import Asistencia, FirmaRegistroClase, RegistroClase
from backend.apps.academico.services.libro_clases_service import LibroClasesService
from backend.apps.auditoria.models import AuditoriaEvento
from backend.apps.cursos.models import Asignatura, Clase
from backend.apps.matriculas.models import Matricula
from tests.common.test_base import BaseTestCase


class ProfesorLibroClasesTest(BaseTestCase):
    """Verifica flujo básico de guardar y firmar registro de clase."""

    def setUp(self):
        super().setUp()
        self.user_profesor = self.crear_usuario_profesor(email='profesor_libro@test.cl', rut='22333444-5')

        self.asignatura = Asignatura.objects.create(
            colegio=self.colegio,
            nombre='Historia',
            codigo='HIS101',
            horas_semanales=3,
            activa=True,
        )

        self.clase = Clase.objects.create(
            colegio=self.colegio,
            curso=self.curso,
            asignatura=self.asignatura,
            profesor=self.user_profesor,
            activo=True,
        )

    def test_servicio_firma_bloquea_edicion(self):
        registro, created = LibroClasesService.upsert_registro_profesor(
            user=self.user_profesor,
            colegio_id=self.colegio.rbd,
            clase_id=self.clase.id,
            fecha=date(2026, 3, 20),
            numero_clase=1,
            contenido_tratado='Edad Media y feudalismo',
            tarea_asignada='Leer capitulo 2',
            observaciones='Buena participacion',
        )
        self.assertTrue(created)

        firmado = LibroClasesService.firmar_registro_profesor(
            user=self.user_profesor,
            colegio_id=self.colegio.rbd,
            registro_id=registro.id_registro,
            ip_address='127.0.0.1',
            user_agent='pytest',
        )

        self.assertTrue(firmado.firmado)
        self.assertIsNotNone(firmado.fecha_firma)
        self.assertEqual(len(firmado.hash_contenido), 64)

        firma = FirmaRegistroClase.objects.get(registro_clase=firmado)
        self.assertEqual(firma.estado, 'FIRMADO')
        self.assertEqual(firma.firma_hash, firmado.hash_contenido)

        with self.assertRaises(ValidationError):
            LibroClasesService.upsert_registro_profesor(
                user=self.user_profesor,
                colegio_id=self.colegio.rbd,
                clase_id=self.clase.id,
                fecha=date(2026, 3, 20),
                numero_clase=1,
                contenido_tratado='Intento de cambio posterior',
                tarea_asignada='Nueva tarea',
                observaciones='No deberia permitir edicion',
            )

    def test_api_guardar_y_firmar_registro(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            return_value=True,
        ):
            response_guardar = self.client.post(
                '/api/profesor/libro-clases/registro/',
                data=json.dumps(
                    {
                        'clase_id': self.clase.id,
                        'fecha': '2026-03-21',
                        'numero_clase': 2,
                        'contenido_tratado': 'Humanismo y Renacimiento',
                        'tarea_asignada': 'Resumen del tema',
                        'observaciones': 'Trabajo en grupos',
                    }
                ),
                content_type='application/json',
            )
            self.assertEqual(response_guardar.status_code, 200)
            payload_guardar = response_guardar.json()
            self.assertTrue(payload_guardar['success'])
            self.assertTrue(payload_guardar['created'])

            registro_id = payload_guardar['registro']['id_registro']

            response_firmar = self.client.post(
                f'/api/profesor/libro-clases/{registro_id}/firmar/',
                data='{}',
                content_type='application/json',
            )
            self.assertEqual(response_firmar.status_code, 200)
            payload_firmar = response_firmar.json()
            self.assertTrue(payload_firmar['success'])
            self.assertTrue(payload_firmar['registro']['firmado'])

            response_guardar_nuevo = self.client.post(
                '/api/profesor/libro-clases/registro/',
                data=json.dumps(
                    {
                        'clase_id': self.clase.id,
                        'fecha': '2026-03-21',
                        'numero_clase': 2,
                        'contenido_tratado': 'Cambio no permitido',
                        'tarea_asignada': 'No aplica',
                        'observaciones': '',
                    }
                ),
                content_type='application/json',
            )
            self.assertEqual(response_guardar_nuevo.status_code, 400)
            payload_error = response_guardar_nuevo.json()
            self.assertFalse(payload_error['success'])

            registro = RegistroClase.objects.get(id_registro=registro_id)
            self.assertEqual(registro.contenido_tratado, 'Humanismo y Renacimiento')

    def test_api_exportar_reporte_superintendencia_csv_pdf(self):
        user_estudiante, _ = self.crear_usuario_estudiante(
            email='est_export@test.cl',
            rut='18888999-1',
        )

        Matricula.objects.create(
            estudiante=user_estudiante,
            colegio=self.colegio,
            curso=self.curso,
            ciclo_academico=self.ciclo,
            estado='ACTIVA',
        )

        Asistencia.objects.create(
            colegio=self.colegio,
            clase=self.clase,
            estudiante=user_estudiante,
            fecha=date(2026, 3, 22),
            estado='P',
            tipo_asistencia='Presencial',
        )

        registro, _ = LibroClasesService.upsert_registro_profesor(
            user=self.user_profesor,
            colegio_id=self.colegio.rbd,
            clase_id=self.clase.id,
            fecha=date(2026, 3, 22),
            numero_clase=1,
            contenido_tratado='Unidad de exportes',
            tarea_asignada='Ninguna',
            observaciones='Reporte de prueba',
        )
        LibroClasesService.firmar_registro_profesor(
            user=self.user_profesor,
            colegio_id=self.colegio.rbd,
            registro_id=registro.id_registro,
            ip_address='127.0.0.1',
            user_agent='pytest',
        )

        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            response_csv = self.client.get('/api/reportes/superintendencia/?format=csv&month=2026-03')
            self.assertEqual(response_csv.status_code, 200)
            self.assertIn('text/csv', response_csv['Content-Type'])
            self.assertIn('.csv', response_csv['Content-Disposition'])
            self.assertIn('superintendencia_decreto67_mensual', response_csv.content.decode('utf-8'))

            response_pdf = self.client.get('/api/reportes/superintendencia/?format=pdf&month=2026-03')
            self.assertEqual(response_pdf.status_code, 200)
            self.assertIn('application/pdf', response_pdf['Content-Type'])
            self.assertIn('.pdf', response_pdf['Content-Disposition'])

    def test_api_exportar_reporte_superintendencia_sin_permiso(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            return_value=False,
        ):
            response = self.client.get('/api/reportes/superintendencia/?format=csv&month=2026-03')
            self.assertEqual(response.status_code, 403)
            payload = response.json()
            self.assertFalse(payload['success'])

        evento = AuditoriaEvento.objects.filter(tabla_afectada='reporte_superintendencia').order_by('-fecha_hora').first()
        self.assertIsNotNone(evento)
        self.assertEqual(evento.accion, AuditoriaEvento.EXPORTAR)
        self.assertEqual(evento.nivel, AuditoriaEvento.NIVEL_WARNING)
        self.assertEqual((evento.metadata or {}).get('result'), 'denied')

    def test_api_exportar_reporte_superintendencia_json_y_xlsx_contrato(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            response_json = self.client.get('/api/reportes/superintendencia/?format=json&month=2026-03')
            self.assertEqual(response_json.status_code, 200)
            payload = response_json.json()
            self.assertTrue(payload['success'])
            self.assertIn('data', payload)
            self.assertEqual(payload['data']['report'], 'superintendencia_decreto67_mensual')
            self.assertEqual(payload['data']['month'], '2026-03')
            self.assertIn('asistencia', payload['data'])
            self.assertIn('matricula', payload['data'])
            self.assertIn('libro_clases', payload['data'])
            self.assertIn('decreto_67', payload['data'])

            response_xlsx = self.client.get('/api/reportes/superintendencia/?format=xlsx&month=2026-03')
            self.assertEqual(response_xlsx.status_code, 200)
            self.assertIn(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                response_xlsx['Content-Type'],
            )
            self.assertIn('.xlsx', response_xlsx['Content-Disposition'])

            workbook = load_workbook(filename=BytesIO(response_xlsx.content))
            self.assertIn('superintendencia_decreto67', workbook.sheetnames)
            ws = workbook['superintendencia_decreto67']
            self.assertEqual(ws['A1'].value, 'reporte')
            self.assertEqual(ws['A2'].value, 'superintendencia_decreto67_mensual')

            response_sige = self.client.get('/api/reportes/superintendencia/?format=sige&month=2026-03')
            self.assertEqual(response_sige.status_code, 200)
            self.assertIn('application/json', response_sige['Content-Type'])
            self.assertIn('_sige.json', response_sige['Content-Disposition'])
            sige_payload = json.loads(response_sige.content.decode('utf-8'))
            self.assertEqual(sige_payload['adapter'], 'sige_ministerial_monthly')
            self.assertEqual(sige_payload['month'], '2026-03')
            self.assertIn('attendance_summary', sige_payload)
            self.assertIn('enrollment_summary', sige_payload)

            evento = AuditoriaEvento.objects.filter(tabla_afectada='reporte_superintendencia').order_by('-fecha_hora').first()
            self.assertIsNotNone(evento)
            self.assertEqual(evento.accion, AuditoriaEvento.EXPORTAR)
            self.assertEqual((evento.metadata or {}).get('result'), 'success')

    def test_api_listar_registros_rbd_coordinador_y_admin(self):
        registro, _ = LibroClasesService.upsert_registro_profesor(
            user=self.user_profesor,
            colegio_id=self.colegio.rbd,
            clase_id=self.clase.id,
            fecha=date(2026, 3, 23),
            numero_clase=1,
            contenido_tratado='Contenido RBD',
            tarea_asignada='Tarea RBD',
            observaciones='Obs RBD',
        )
        self.assertIsNotNone(registro.id_registro)

        rol_coord = Role.objects.get_or_create(nombre='COORDINADOR_ACADEMICO')[0]
        user_coord = User.objects.create_user(
            email='coord_rbd@test.cl',
            password='test123456',
            nombre='Coord',
            apellido_paterno='Academico',
            role=rol_coord,
            rbd_colegio=self.colegio.rbd,
        )

        user_admin = self.crear_usuario_admin(email='admin_rbd@test.cl', rut='14444555-6')

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'LIBRO_CLASE_VIEW_RBD',
        ):
            self.client.force_login(user_coord)
            response_coord = self.client.get('/api/coordinador/libro-clases/?fecha=2026-03-23')
            self.assertEqual(response_coord.status_code, 200)
            payload_coord = response_coord.json()
            self.assertTrue(payload_coord['success'])
            self.assertGreaterEqual(len(payload_coord['registros']), 1)

            self.client.force_login(user_admin)
            response_admin = self.client.get('/api/admin-escolar/libro-clases/?fecha=2026-03-23')
            self.assertEqual(response_admin.status_code, 200)
            payload_admin = response_admin.json()
            self.assertTrue(payload_admin['success'])
            self.assertGreaterEqual(len(payload_admin['registros']), 1)

    def test_api_auditoria_superintendencia_listado_y_filtro(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            # Genera eventos de auditoria exitosos via exportacion.
            response_export_1 = self.client.get('/api/reportes/superintendencia/?format=csv&month=2026-03')
            self.assertEqual(response_export_1.status_code, 200)
            response_export_2 = self.client.get('/api/reportes/superintendencia/?format=pdf&month=2026-03')
            self.assertEqual(response_export_2.status_code, 200)

            response_list = self.client.get('/api/reportes/superintendencia/auditoria/?month=2026-03&page=1&page_size=1')
            self.assertEqual(response_list.status_code, 200)
            payload = response_list.json()
            self.assertTrue(payload['success'])
            self.assertGreaterEqual(payload['total'], 1)
            self.assertEqual(payload['page'], 1)
            self.assertEqual(payload['page_size'], 1)
            self.assertEqual(len(payload['eventos']), 1)
            self.assertGreaterEqual(payload['total_pages'], 1)
            self.assertTrue(any(item.get('month') == '2026-03' for item in payload['eventos']))

            response_list_page_2 = self.client.get('/api/reportes/superintendencia/auditoria/?month=2026-03&page=2&page_size=1')
            self.assertEqual(response_list_page_2.status_code, 200)
            payload_page_2 = response_list_page_2.json()
            self.assertTrue(payload_page_2['success'])
            self.assertEqual(payload_page_2['page'], 2)

    def test_api_auditoria_superintendencia_descarga_csv(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            response_export = self.client.get('/api/reportes/superintendencia/?format=csv&month=2026-03')
            self.assertEqual(response_export.status_code, 200)

            response_csv = self.client.get('/api/reportes/superintendencia/auditoria/?month=2026-03&download=csv')
            self.assertEqual(response_csv.status_code, 200)
            self.assertIn('text/csv', response_csv['Content-Type'])
            self.assertIn('auditoria_superintendencia_', response_csv['Content-Disposition'])

            csv_text = response_csv.content.decode('utf-8')
            self.assertIn('id,fecha_hora,usuario_id,usuario_nombre,accion,categoria,nivel,format,month,result,filename,descripcion', csv_text)
            self.assertIn('2026-03', csv_text)

    def test_api_auditoria_superintendencia_filtro_usuario_y_fechas(self):
        self.client.force_login(self.user_profesor)

        user_otro = self.crear_usuario_profesor(email='otro_auditoria@test.cl', rut='23444555-6')
        AuditoriaEvento.objects.create(
            usuario=self.user_profesor,
            colegio_rbd=str(self.colegio.rbd),
            accion=AuditoriaEvento.EXPORTAR,
            tabla_afectada='reporte_superintendencia',
            descripcion='Evento profesor principal',
            categoria=AuditoriaEvento.CATEGORIA_ACADEMICO,
            nivel=AuditoriaEvento.NIVEL_INFO,
            metadata={'format': 'csv', 'month': '2026-03', 'result': 'success'},
        )
        AuditoriaEvento.objects.create(
            usuario=user_otro,
            colegio_rbd=str(self.colegio.rbd),
            accion=AuditoriaEvento.EXPORTAR,
            tabla_afectada='reporte_superintendencia',
            descripcion='Evento otro usuario',
            categoria=AuditoriaEvento.CATEGORIA_ACADEMICO,
            nivel=AuditoriaEvento.NIVEL_INFO,
            metadata={'format': 'pdf', 'month': '2026-03', 'result': 'success'},
        )

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            response = self.client.get(
                f'/api/reportes/superintendencia/auditoria/?usuario={self.user_profesor.id}&fecha_desde=2000-01-01&fecha_hasta=2100-12-31'
            )
            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertTrue(payload['success'])
            self.assertGreaterEqual(payload['total'], 1)
            self.assertTrue(all(item['usuario_id'] == self.user_profesor.id for item in payload['eventos']))

    def test_api_auditoria_superintendencia_parametro_fecha_invalido(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            response = self.client.get('/api/reportes/superintendencia/auditoria/?fecha_desde=2026-99-99')
            self.assertEqual(response.status_code, 400)
            payload = response.json()
            self.assertFalse(payload['success'])

    def test_api_auditoria_superintendencia_ordenamiento_resultado(self):
        self.client.force_login(self.user_profesor)

        AuditoriaEvento.objects.create(
            usuario=self.user_profesor,
            colegio_rbd=str(self.colegio.rbd),
            accion=AuditoriaEvento.EXPORTAR,
            tabla_afectada='reporte_superintendencia',
            descripcion='Evento success',
            categoria=AuditoriaEvento.CATEGORIA_ACADEMICO,
            nivel=AuditoriaEvento.NIVEL_INFO,
            metadata={'format': 'csv', 'month': '2026-03', 'result': 'success'},
        )
        AuditoriaEvento.objects.create(
            usuario=self.user_profesor,
            colegio_rbd=str(self.colegio.rbd),
            accion=AuditoriaEvento.EXPORTAR,
            tabla_afectada='reporte_superintendencia',
            descripcion='Evento denied',
            categoria=AuditoriaEvento.CATEGORIA_ACADEMICO,
            nivel=AuditoriaEvento.NIVEL_WARNING,
            metadata={'format': 'pdf', 'month': '2026-03', 'result': 'denied'},
        )

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            response = self.client.get(
                '/api/reportes/superintendencia/auditoria/?month=2026-03&sort_by=resultado&sort_dir=asc&page_size=50'
            )
            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertTrue(payload['success'])
            self.assertEqual(payload['sort_by'], 'resultado')
            self.assertEqual(payload['sort_dir'], 'asc')

            resultados = [item.get('result') for item in payload['eventos'] if item.get('result') in {'denied', 'success'}]
            self.assertIn('denied', resultados)
            self.assertIn('success', resultados)
            self.assertLessEqual(resultados.index('denied'), resultados.index('success'))

    def test_api_auditoria_superintendencia_parametro_sort_invalido(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            side_effect=lambda _u, cap, school_id=None: cap == 'REPORT_EXPORT_SUPERINTENDENCIA',
        ):
            response = self.client.get('/api/reportes/superintendencia/auditoria/?sort_by=inexistente')
            self.assertEqual(response.status_code, 400)
            payload = response.json()
            self.assertFalse(payload['success'])

    def test_api_auditoria_superintendencia_sin_permiso(self):
        self.client.force_login(self.user_profesor)

        with patch(
            'backend.apps.core.views.profesor.libro_clases_api.PolicyService.has_capability',
            return_value=False,
        ):
            response = self.client.get('/api/reportes/superintendencia/auditoria/?month=2026-03')
            self.assertEqual(response.status_code, 403)
            payload = response.json()
            self.assertFalse(payload['success'])
