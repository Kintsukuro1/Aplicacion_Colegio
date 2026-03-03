"""
Utilidades para exportar reportes académicos a PDF y Excel.

Funcionalidades:
- Generación de PDF con reportlab
- Generación de Excel con openpyxl
- Formato profesional con logos, headers y estilos
"""
import io
from datetime import date
from typing import Dict, List

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PDFReportExporter:
    """Exportador de reportes académicos a PDF"""
    
    @staticmethod
    def generate_student_academic_pdf(reporte_data: Dict, colegio) -> HttpResponse:
        """
        Genera PDF de reporte académico de estudiante.
        
        Args:
            reporte_data: Diccionario con datos del reporte
            colegio: Instancia del colegio
            
        Returns:
            HttpResponse con PDF adjunto
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a472a'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2e7d32'),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_LEFT
        )
        
        # Header institucional
        story.append(Paragraph(f"<b>{colegio.nombre_institucion}</b>", title_style))
        story.append(Paragraph(f"RBD: {colegio.rbd}", subtitle_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Título del reporte
        story.append(Paragraph("<b>REPORTE ACADÉMICO DEL ESTUDIANTE</b>", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Información del estudiante
        estudiante = reporte_data['estudiante']
        info_estudiante = [
            ['<b>Nombre Completo:</b>', f"{estudiante['nombre']} {estudiante['apellido_paterno']} {estudiante['apellido_materno']}"],
            ['<b>RUT:</b>', estudiante['rut']],
            ['<b>Curso:</b>', reporte_data['curso']],
            ['<b>Período:</b>', reporte_data['periodo']],
            ['<b>Fecha de Generación:</b>', reporte_data['fecha_generacion'].strftime('%d/%m/%Y')],
        ]
        
        info_table = Table(info_estudiante, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1a472a')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Resumen académico
        resumen_data = [
            ['<b>Promedio General:</b>', f"{reporte_data['promedio_general']:.1f}"],
            ['<b>Estado:</b>', 'Aprobado' if reporte_data['promedio_general'] >= 4.0 else 'Reprobado'],
        ]
        
        resumen_table = Table(resumen_data, colWidths=[2*inch, 4*inch])
        resumen_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1a472a')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e8f5e9')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#2e7d32')),
        ]))
        story.append(resumen_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Tabla de calificaciones por asignatura
        story.append(Paragraph("<b>CALIFICACIONES POR ASIGNATURA</b>", subtitle_style))
        story.append(Spacer(1, 0.1*inch))
        
        calificaciones_data = [['Asignatura', 'Nota Final', 'Estado']]
        for asignatura in reporte_data['asignaturas']:
            color_nota = 'green' if asignatura['nota_final'] >= 4.0 else 'red'
            calificaciones_data.append([
                asignatura['asignatura'],
                f"{asignatura['nota_final']:.1f}",
                asignatura['estado']
            ])
        
        calificaciones_table = Table(calificaciones_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        calificaciones_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(calificaciones_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Estadísticas de asistencia
        story.append(Paragraph("<b>ESTADÍSTICAS DE ASISTENCIA</b>", subtitle_style))
        story.append(Spacer(1, 0.1*inch))
        
        asistencia = reporte_data['asistencia']
        asistencia_data = [
            ['Total de Clases', 'Presencias', 'Porcentaje de Asistencia'],
            [str(asistencia['total_clases']), str(asistencia['presentes']), f"{asistencia['porcentaje']:.1f}%"]
        ]
        
        asistencia_table = Table(asistencia_data, colWidths=[2*inch, 2*inch, 2*inch])
        asistencia_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#e3f2fd')),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(asistencia_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            f"Documento generado electrónicamente el {date.today().strftime('%d/%m/%Y')} - {colegio.nombre_institucion}",
            footer_style
        ))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Preparar response
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"reporte_academico_{estudiante['rut']}_{date.today().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class ExcelReportExporter:
    """Exportador de reportes académicos a Excel"""
    
    @staticmethod
    def generate_student_academic_excel(reporte_data: Dict, colegio) -> HttpResponse:
        """
        Genera Excel de reporte académico de estudiante con múltiples hojas.
        
        Args:
            reporte_data: Diccionario con datos del reporte
            colegio: Instancia del colegio
            
        Returns:
            HttpResponse con Excel adjunto
        """
        wb = Workbook()
        
        # Estilos reutilizables
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2e7d32', end_color='2e7d32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        title_font = Font(name='Calibri', size=14, bold=True, color='1a472a')
        subtitle_font = Font(name='Calibri', size=11, bold=True, color='2e7d32')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ===== HOJA 1: RESUMEN =====
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        estudiante = reporte_data['estudiante']
        
        # Header institucional
        ws_resumen['A1'] = colegio.nombre_institucion
        ws_resumen['A1'].font = Font(name='Calibri', size=16, bold=True, color='1a472a')
        ws_resumen['A1'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A1:D1')
        
        ws_resumen['A2'] = f"RBD: {colegio.rbd}"
        ws_resumen['A2'].font = Font(name='Calibri', size=10, color='666666')
        ws_resumen['A2'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A2:D2')
        
        ws_resumen['A4'] = "REPORTE ACADÉMICO DEL ESTUDIANTE"
        ws_resumen['A4'].font = Font(name='Calibri', size=14, bold=True)
        ws_resumen['A4'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A4:D4')
        
        # Información del estudiante
        row = 6
        ws_resumen[f'A{row}'] = "Nombre Completo:"
        ws_resumen[f'A{row}'].font = subtitle_font
        ws_resumen[f'B{row}'] = f"{estudiante['nombre']} {estudiante['apellido_paterno']} {estudiante['apellido_materno']}"
        ws_resumen.merge_cells(f'B{row}:D{row}')
        
        row += 1
        ws_resumen[f'A{row}'] = "RUT:"
        ws_resumen[f'A{row}'].font = subtitle_font
        ws_resumen[f'B{row}'] = estudiante['rut']
        
        row += 1
        ws_resumen[f'A{row}'] = "Curso:"
        ws_resumen[f'A{row}'].font = subtitle_font
        ws_resumen[f'B{row}'] = reporte_data['curso']
        
        row += 1
        ws_resumen[f'A{row}'] = "Período:"
        ws_resumen[f'A{row}'].font = subtitle_font
        ws_resumen[f'B{row}'] = reporte_data['periodo']
        
        row += 1
        ws_resumen[f'A{row}'] = "Fecha de Generación:"
        ws_resumen[f'A{row}'].font = subtitle_font
        ws_resumen[f'B{row}'] = reporte_data['fecha_generacion'].strftime('%d/%m/%Y')
        
        # Resumen académico destacado
        row += 3
        ws_resumen[f'A{row}'] = "PROMEDIO GENERAL"
        ws_resumen[f'A{row}'].font = header_font
        ws_resumen[f'A{row}'].fill = header_fill
        ws_resumen[f'A{row}'].alignment = header_alignment
        
        ws_resumen[f'B{row}'] = "ESTADO"
        ws_resumen[f'B{row}'].font = header_font
        ws_resumen[f'B{row}'].fill = header_fill
        ws_resumen[f'B{row}'].alignment = header_alignment
        
        row += 1
        promedio = reporte_data['promedio_general']
        ws_resumen[f'A{row}'] = f"{promedio:.1f}"
        ws_resumen[f'A{row}'].font = Font(name='Calibri', size=18, bold=True)
        ws_resumen[f'A{row}'].alignment = Alignment(horizontal='center')
        
        estado = 'Aprobado' if promedio >= 4.0 else 'Reprobado'
        ws_resumen[f'B{row}'] = estado
        ws_resumen[f'B{row}'].font = Font(name='Calibri', size=14, bold=True, color='00b050' if promedio >= 4.0 else 'ff0000')
        ws_resumen[f'B{row}'].alignment = Alignment(horizontal='center')
        
        # Asistencia
        row += 3
        ws_resumen[f'A{row}'] = "ASISTENCIA"
        ws_resumen[f'A{row}'].font = subtitle_font
        ws_resumen.merge_cells(f'A{row}:D{row}')
        
        row += 1
        asistencia_headers = ['Total Clases', 'Presencias', 'Porcentaje']
        for idx, header in enumerate(asistencia_headers, start=1):
            cell = ws_resumen.cell(row=row, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color='1565c0', end_color='1565c0', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row += 1
        asistencia = reporte_data['asistencia']
        ws_resumen[f'A{row}'] = asistencia['total_clases']
        ws_resumen[f'A{row}'].alignment = Alignment(horizontal='center')
        ws_resumen[f'A{row}'].border = thin_border
        
        ws_resumen[f'B{row}'] = asistencia['presentes']
        ws_resumen[f'B{row}'].alignment = Alignment(horizontal='center')
        ws_resumen[f'B{row}'].border = thin_border
        
        ws_resumen[f'C{row}'] = f"{asistencia['porcentaje']:.1f}%"
        ws_resumen[f'C{row}'].alignment = Alignment(horizontal='center')
        ws_resumen[f'C{row}'].border = thin_border
        
        # Ajustar anchos de columna
        ws_resumen.column_dimensions['A'].width = 25
        ws_resumen.column_dimensions['B'].width = 25
        ws_resumen.column_dimensions['C'].width = 20
        ws_resumen.column_dimensions['D'].width = 20
        
        # ===== HOJA 2: CALIFICACIONES DETALLE =====
        ws_calificaciones = wb.create_sheet(title="Calificaciones")
        
        # Header
        ws_calificaciones['A1'] = "CALIFICACIONES POR ASIGNATURA"
        ws_calificaciones['A1'].font = title_font
        ws_calificaciones['A1'].alignment = Alignment(horizontal='center')
        ws_calificaciones.merge_cells('A1:C1')
        
        # Headers de tabla
        headers = ['Asignatura', 'Nota Final', 'Estado']
        for idx, header in enumerate(headers, start=1):
            cell = ws_calificaciones.cell(row=3, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos de calificaciones
        row = 4
        for asignatura_data in reporte_data['asignaturas']:
            ws_calificaciones[f'A{row}'] = asignatura_data['asignatura']
            ws_calificaciones[f'A{row}'].border = thin_border
            
            nota = asignatura_data['nota_final']
            ws_calificaciones[f'B{row}'] = f"{nota:.1f}"
            ws_calificaciones[f'B{row}'].alignment = Alignment(horizontal='center')
            ws_calificaciones[f'B{row}'].border = thin_border
            
            # Color según nota
            if nota >= 6.0:
                fill_color = 'c6efce'  # Verde claro
            elif nota >= 4.0:
                fill_color = 'ffeb9c'  # Amarillo claro
            else:
                fill_color = 'ffc7ce'  # Rojo claro
            ws_calificaciones[f'B{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            ws_calificaciones[f'C{row}'] = asignatura_data['estado']
            ws_calificaciones[f'C{row}'].alignment = Alignment(horizontal='center')
            ws_calificaciones[f'C{row}'].border = thin_border
            ws_calificaciones[f'C{row}'].font = Font(
                bold=True,
                color='00b050' if nota >= 4.0 else 'ff0000'
            )
            
            row += 1
        
        # Ajustar anchos
        ws_calificaciones.column_dimensions['A'].width = 35
        ws_calificaciones.column_dimensions['B'].width = 15
        ws_calificaciones.column_dimensions['C'].width = 15
        
        # ===== HOJA 3: ESTADÍSTICAS =====
        ws_estadisticas = wb.create_sheet(title="Estadísticas")
        
        ws_estadisticas['A1'] = "ESTADÍSTICAS ACADÉMICAS"
        ws_estadisticas['A1'].font = title_font
        ws_estadisticas['A1'].alignment = Alignment(horizontal='center')
        ws_estadisticas.merge_cells('A1:B1')
        
        # Calcular estadísticas
        notas = [a['nota_final'] for a in reporte_data['asignaturas'] if a['nota_final'] > 0]
        
        stats_data = [
            ('Total de Asignaturas', len(notas)),
            ('Promedio General', f"{promedio:.2f}"),
            ('Nota Más Alta', f"{max(notas):.1f}" if notas else "N/A"),
            ('Nota Más Baja', f"{min(notas):.1f}" if notas else "N/A"),
            ('Asignaturas Aprobadas', sum(1 for n in notas if n >= 4.0)),
            ('Asignaturas Reprobadas', sum(1 for n in notas if n < 4.0)),
        ]
        
        row = 3
        for label, value in stats_data:
            ws_estadisticas[f'A{row}'] = label
            ws_estadisticas[f'A{row}'].font = subtitle_font
            ws_estadisticas[f'A{row}'].border = thin_border
            
            ws_estadisticas[f'B{row}'] = value
            ws_estadisticas[f'B{row}'].alignment = Alignment(horizontal='center')
            ws_estadisticas[f'B{row}'].border = thin_border
            
            row += 1
        
        ws_estadisticas.column_dimensions['A'].width = 30
        ws_estadisticas.column_dimensions['B'].width = 20
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Preparar response
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reporte_academico_{estudiante['rut']}_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
